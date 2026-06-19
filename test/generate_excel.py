"""
Generate test_1.xlsx from auto_results.json
"""
import json
import sys
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

TEST_DIR = Path(__file__).parent
RESULTS_FILE = TEST_DIR / "auto_results.json"
OUTPUT_FILE = TEST_DIR / "test_1.xlsx"

# Domain mapping based on Q-ID ranges
DOMAIN_MAP = {
    range(1, 26):   "Parking",
    range(26, 41):  "Asset / Building",
    range(41, 66):  "Telemetry",
    range(66, 76):  "DMP Status",
    range(76, 96):  "ISO 37122",
    range(96, 101): "Cross-domain",
    range(101, 120):"SQL Pairs Bổ Sung",
    range(120, 150):"AC.xlsx / Multi-turn",
}

def get_domain(q_id: str) -> str:
    num = int(q_id[1:])
    for r, domain in DOMAIN_MAP.items():
        if num in r:
            return domain
    return "Unknown"

def get_batch(q_id: str) -> str:
    num = int(q_id[1:])
    batch_num = (num - 1) // 10 + 1
    return f"Batch {batch_num:02d}"

# ── Style helpers ──────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE_H = PatternFill("solid", fgColor="2F5496")
GRAY_H = PatternFill("solid", fgColor="D9E1F2")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")

WHITE_BOLD = Font(bold=True, color="FFFFFF")
DARK_BOLD  = Font(bold=True)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = BLUE_H
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = thin_border()
    return c

def apply_border(ws, row, col):
    ws.cell(row=row, column=col).border = thin_border()

# ── Load data ──────────────────────────────────────────────────────────────────
with open(RESULTS_FILE, encoding="utf-8") as f:
    results = json.load(f)

items = []
for qid, r in sorted(results.items()):
    items.append({
        "id":       r["id"],
        "batch":    get_batch(r["id"]),
        "domain":   get_domain(r["id"]),
        "question": r["question"],
        "verdict":  r["verdict"],
        "type":     r.get("type", ""),
        "status":   r.get("status", ""),
        "note":     r.get("note", ""),
        "elapsed":  r.get("elapsed", 0),
        "sql_len":  len(r.get("sql", "") or ""),
        "timestamp": r.get("timestamp", ""),
    })

total  = len(items)
passed = sum(1 for x in items if x["verdict"] == "PASS")
failed = sum(1 for x in items if x["verdict"] == "FAIL")
pct    = passed / total * 100

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Tất cả kết quả
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Tất cả kết quả"

# Title row
ws1.merge_cells("A1:J1")
t = ws1["A1"]
t.value = f"Wren AI — Test Report   |   Ngày: {datetime.now():%Y-%m-%d}   |   Tổng: {total}   PASS: {passed} ({pct:.1f}%)   FAIL: {failed}"
t.font  = Font(bold=True, size=13, color="FFFFFF")
t.fill  = BLUE_H
t.alignment = CENTER
ws1.row_dimensions[1].height = 28

# Header
cols = ["ID", "Batch", "Domain", "Câu hỏi", "Kết quả", "Type", "Status", "Ghi chú", "Thời gian (s)", "SQL length"]
for ci, col in enumerate(cols, 1):
    header_cell(ws1, 2, ci, col)
ws1.row_dimensions[2].height = 20

# Column widths
widths = [7, 9, 20, 60, 9, 14, 12, 55, 13, 11]
for ci, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# Data rows
for ri, item in enumerate(items, 3):
    row_fill = GREEN if item["verdict"] == "PASS" else RED
    ws1.cell(ri, 1, item["id"]).alignment        = CENTER
    ws1.cell(ri, 2, item["batch"]).alignment     = CENTER
    ws1.cell(ri, 3, item["domain"]).alignment    = LEFT_WRAP
    ws1.cell(ri, 4, item["question"]).alignment  = LEFT_WRAP
    vdict = ws1.cell(ri, 5, item["verdict"])
    vdict.alignment = CENTER
    vdict.fill      = row_fill
    vdict.font      = DARK_BOLD
    ws1.cell(ri, 6, item["type"]).alignment      = CENTER
    ws1.cell(ri, 7, item["status"]).alignment    = CENTER
    ws1.cell(ri, 8, item["note"]).alignment      = LEFT_WRAP
    elap = round(item["elapsed"], 1) if item["elapsed"] else ""
    ws1.cell(ri, 9, elap).alignment              = CENTER
    ws1.cell(ri, 10, item["sql_len"] or "").alignment = CENTER

    bg = LIGHT_GRAY if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for ci in range(1, 11):
        c = ws1.cell(ri, ci)
        c.border = thin_border()
        if ci != 5:
            c.fill = bg
    ws1.row_dimensions[ri].height = 40

ws1.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Thống kê tổng quan
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Thống kê")

def section_title(ws, row, title, ncols=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = BLUE_H
    c.alignment = CENTER
    ws.row_dimensions[row].height = 22

# ── A. Tổng quan ─────────────────────────────────────────────────────────────
section_title(ws2, 1, "A. Tổng quan kết quả")
for ci, h in enumerate(["Chỉ số", "Giá trị", "Tỉ lệ"], 1):
    header_cell(ws2, 2, ci, h)

summary_rows = [
    ("Tổng câu hỏi",       total,  "100%"),
    ("PASS (sinh SQL đúng)", passed, f"{pct:.1f}%"),
    ("FAIL",                failed, f"{100-pct:.1f}%"),
    ("- FAIL do GENERAL",   sum(1 for x in items if x["verdict"]=="FAIL" and x["type"]=="GENERAL"),   ""),
    ("- FAIL do Schema lỗi",sum(1 for x in items if x["verdict"]=="FAIL" and x["status"]=="FAILED"),  ""),
    ("Avg response time (s)", round(sum(x["elapsed"] for x in items if x["elapsed"])/total,1), ""),
    ("Max response time (s)", round(max(x["elapsed"] for x in items if x["elapsed"]),1), ""),
]
for ri, (label, val, pct_val) in enumerate(summary_rows, 3):
    ws2.cell(ri, 1, label).font = DARK_BOLD
    ws2.cell(ri, 2, val).alignment = CENTER
    ws2.cell(ri, 3, pct_val).alignment = CENTER
    fill = LIGHT_GRAY if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for ci in range(1, 4):
        ws2.cell(ri, ci).border = thin_border()
        ws2.cell(ri, ci).fill = fill

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12

# ── B. Theo Batch ──────────────────────────────────────────────────────────────
row_start_batch = 12
section_title(ws2, row_start_batch, "B. Kết quả theo Batch", 5)
for ci, h in enumerate(["Batch", "Tổng", "PASS", "FAIL", "Tỉ lệ PASS"], 1):
    header_cell(ws2, row_start_batch + 1, ci, h)

batch_stats = {}
for item in items:
    b = item["batch"]
    batch_stats.setdefault(b, {"total": 0, "pass": 0, "fail": 0})
    batch_stats[b]["total"] += 1
    if item["verdict"] == "PASS":
        batch_stats[b]["pass"] += 1
    else:
        batch_stats[b]["fail"] += 1

for ri, (batch, s) in enumerate(sorted(batch_stats.items()), row_start_batch + 2):
    p_pct = s["pass"] / s["total"] * 100
    ws2.cell(ri, 1, batch).alignment = CENTER
    ws2.cell(ri, 2, s["total"]).alignment = CENTER
    ws2.cell(ri, 3, s["pass"]).alignment = CENTER
    ws2.cell(ri, 4, s["fail"]).alignment = CENTER
    pct_cell = ws2.cell(ri, 5, f"{p_pct:.0f}%")
    pct_cell.alignment = CENTER
    pct_cell.fill = GREEN if p_pct >= 70 else YELLOW if p_pct >= 50 else RED
    fill = LIGHT_GRAY if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for ci in range(1, 6):
        ws2.cell(ri, ci).border = thin_border()
        if ci != 5:
            ws2.cell(ri, ci).fill = fill

for ci in range(1, 6):
    ws2.column_dimensions[get_column_letter(ci)].width = max(
        ws2.column_dimensions[get_column_letter(ci)].width or 0, 12
    )

# ── C. Theo Domain ─────────────────────────────────────────────────────────────
row_start_domain = row_start_batch + 2 + len(batch_stats) + 2
section_title(ws2, row_start_domain, "C. Kết quả theo Domain", 5)
for ci, h in enumerate(["Domain", "Tổng", "PASS", "FAIL", "Tỉ lệ PASS"], 1):
    header_cell(ws2, row_start_domain + 1, ci, h)

domain_stats = {}
for item in items:
    d = item["domain"]
    domain_stats.setdefault(d, {"total": 0, "pass": 0, "fail": 0})
    domain_stats[d]["total"] += 1
    if item["verdict"] == "PASS":
        domain_stats[d]["pass"] += 1
    else:
        domain_stats[d]["fail"] += 1

for ri, (domain, s) in enumerate(domain_stats.items(), row_start_domain + 2):
    p_pct = s["pass"] / s["total"] * 100
    ws2.cell(ri, 1, domain).alignment = LEFT_WRAP
    ws2.cell(ri, 2, s["total"]).alignment = CENTER
    ws2.cell(ri, 3, s["pass"]).alignment = CENTER
    ws2.cell(ri, 4, s["fail"]).alignment = CENTER
    pct_cell = ws2.cell(ri, 5, f"{p_pct:.0f}%")
    pct_cell.alignment = CENTER
    pct_cell.fill = GREEN if p_pct >= 70 else YELLOW if p_pct >= 50 else RED
    fill = LIGHT_GRAY if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for ci in range(1, 6):
        ws2.cell(ri, ci).border = thin_border()
        if ci != 5:
            ws2.cell(ri, ci).fill = fill

# ── D. Loại lỗi FAIL ──────────────────────────────────────────────────────────
row_start_err = row_start_domain + 2 + len(domain_stats) + 2
section_title(ws2, row_start_err, "D. Phân loại nguyên nhân FAIL", 4)
for ci, h in enumerate(["Loại lỗi", "Số câu", "Tỉ lệ / FAIL", "Câu ID"], 1):
    header_cell(ws2, row_start_err + 1, ci, h)

fail_general = [x for x in items if x["verdict"]=="FAIL" and x["type"]=="GENERAL"]
fail_schema  = [x for x in items if x["verdict"]=="FAIL" and x["status"]=="FAILED"]

err_rows = [
    ("GENERAL (không sinh SQL)", len(fail_general),
     f"{len(fail_general)/failed*100:.0f}%",
     ", ".join(x["id"] for x in fail_general)),
    ("NO_RELEVANT_SQL (schema lỗi)", len(fail_schema),
     f"{len(fail_schema)/failed*100:.0f}%",
     ", ".join(x["id"] for x in fail_schema)),
]
for ri, (label, cnt, pct_val, ids) in enumerate(err_rows, row_start_err + 2):
    ws2.cell(ri, 1, label).font = DARK_BOLD
    ws2.cell(ri, 2, cnt).alignment = CENTER
    ws2.cell(ri, 3, pct_val).alignment = CENTER
    id_cell = ws2.cell(ri, 4, ids)
    id_cell.alignment = LEFT_WRAP
    id_cell.fill = RED
    for ci in range(1, 5):
        ws2.cell(ri, ci).border = thin_border()

ws2.column_dimensions["D"].width = 80

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — FAIL questions chi tiết
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("FAIL chi tiết")
ws3.merge_cells("A1:G1")
t3 = ws3["A1"]
t3.value = f"FAIL Questions — {failed} câu   ({failed/total*100:.1f}% tổng số)"
t3.font  = Font(bold=True, size=12, color="FFFFFF")
t3.fill  = PatternFill("solid", fgColor="C00000")
t3.alignment = CENTER
ws3.row_dimensions[1].height = 24

fail_cols = ["ID", "Batch", "Domain", "Câu hỏi", "Type", "Status", "Ghi chú / Lỗi"]
for ci, col in enumerate(fail_cols, 1):
    header_cell(ws3, 2, ci, col)

fail_items = [x for x in items if x["verdict"] == "FAIL"]
for ri, item in enumerate(fail_items, 3):
    ws3.cell(ri, 1, item["id"]).alignment  = CENTER
    ws3.cell(ri, 2, item["batch"]).alignment = CENTER
    ws3.cell(ri, 3, item["domain"]).alignment = LEFT_WRAP
    ws3.cell(ri, 4, item["question"]).alignment = LEFT_WRAP
    ws3.cell(ri, 5, item["type"]).alignment  = CENTER
    ws3.cell(ri, 6, item["status"]).alignment = CENTER
    note_cell = ws3.cell(ri, 7, item["note"])
    note_cell.alignment = LEFT_WRAP
    fill = RED if item["status"] == "FAILED" else YELLOW
    for ci in range(1, 8):
        ws3.cell(ri, ci).border = thin_border()
        ws3.cell(ri, ci).fill = fill
    ws3.row_dimensions[ri].height = 38

ws3_widths = [7, 9, 20, 60, 16, 12, 65]
for ci, w in enumerate(ws3_widths, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

ws3.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Dashboard nhanh
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Dashboard")
ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 14

# KPI cards
ws4.merge_cells("A1:B1")
kpi_title = ws4["A1"]
kpi_title.value = "KPI — Wren AI Test"
kpi_title.font  = Font(bold=True, size=14, color="FFFFFF")
kpi_title.fill  = BLUE_H
kpi_title.alignment = CENTER
ws4.row_dimensions[1].height = 30

kpis = [
    ("Tổng câu test", total),
    ("PASS", passed),
    ("FAIL", failed),
    ("Tỉ lệ PASS", f"{pct:.1f}%"),
    ("Tỉ lệ FAIL", f"{100-pct:.1f}%"),
    ("FAIL do GENERAL", sum(1 for x in items if x["verdict"]=="FAIL" and x["type"]=="GENERAL")),
    ("FAIL do Schema lỗi", sum(1 for x in items if x["verdict"]=="FAIL" and x["status"]=="FAILED")),
    ("Avg response (s)", f"{sum(x['elapsed'] for x in items if x['elapsed'])/total:.1f}"),
    ("Ngày test", datetime.now().strftime("%Y-%m-%d")),
]
for ri, (label, val) in enumerate(kpis, 2):
    lc = ws4.cell(ri, 1, label)
    lc.font = DARK_BOLD
    lc.fill = GRAY_H
    lc.border = thin_border()
    lc.alignment = LEFT_WRAP
    vc = ws4.cell(ri, 2, val)
    vc.alignment = CENTER
    vc.border = thin_border()
    if label == "PASS":
        vc.fill = GREEN
    elif label in ("FAIL", "Tỉ lệ FAIL"):
        vc.fill = RED
    elif label == "Tỉ lệ PASS":
        vc.fill = GREEN
    ws4.row_dimensions[ri].height = 22

# Pie chart data (col D-E)
ws4.cell(1, 4, "Kết quả").font = DARK_BOLD
ws4.cell(2, 4, "PASS"); ws4.cell(2, 5, passed)
ws4.cell(3, 4, "FAIL"); ws4.cell(3, 5, failed)

pie = PieChart()
pie.title = "PASS vs FAIL"
pie.style = 10
data_ref   = Reference(ws4, min_col=5, min_row=2, max_row=3)
labels_ref = Reference(ws4, min_col=4, min_row=2, max_row=3)
pie.add_data(data_ref)
pie.set_categories(labels_ref)
pie.dataLabels = openpyxl.chart.label.DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
# Color slices
slice_pass = DataPoint(idx=0)
slice_pass.graphicalProperties.solidFill = "70AD47"
slice_fail = DataPoint(idx=1)
slice_fail.graphicalProperties.solidFill = "FF0000"
pie.series[0].dPt = [slice_pass, slice_fail]
pie.width  = 14
pie.height = 12
ws4.add_chart(pie, "D4")

# Bar chart — batch pass rate
ws4.cell(20, 4, "Batch").font = DARK_BOLD
ws4.cell(20, 5, "PASS").font  = DARK_BOLD
ws4.cell(20, 6, "FAIL").font  = DARK_BOLD
for i, (batch, s) in enumerate(sorted(batch_stats.items()), 21):
    ws4.cell(i, 4, batch)
    ws4.cell(i, 5, s["pass"])
    ws4.cell(i, 6, s["fail"])

bar = BarChart()
bar.type    = "col"
bar.title   = "PASS / FAIL theo Batch"
bar.y_axis.title = "Số câu"
bar.x_axis.title = "Batch"
bar.style   = 10
data_b   = Reference(ws4, min_col=5, max_col=6, min_row=20, max_row=20+len(batch_stats))
cats_b   = Reference(ws4, min_col=4, min_row=21, max_row=20+len(batch_stats))
bar.add_data(data_b, titles_from_data=True)
bar.set_categories(cats_b)
bar.series[0].graphicalProperties.solidFill = "70AD47"
bar.series[1].graphicalProperties.solidFill = "FF0000"
bar.width  = 22
bar.height = 12
ws4.add_chart(bar, "D20")

wb.save(OUTPUT_FILE)
print(f"Done! Saved to: {OUTPUT_FILE}")
print(f"Total: {total} | PASS: {passed} ({pct:.1f}%) | FAIL: {failed}")
